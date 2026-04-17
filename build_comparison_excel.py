"""
build_comparison_excel.py
-------------------------
Genera el archivo Excel de ficha técnica comparativa MRI & CT.

Estructura:
  - Sheet "MRI_1.5T"      — comparativa 4 fabricantes, 17 parámetros
  - Sheet "MRI_3T"        — comparativa 3 fabricantes, 17 parámetros
  - Sheet "CT_64slices"   — comparativa 4 fabricantes, 15 parámetros
  - Sheet "CT_128slices"  — comparativa 3 fabricantes, 15 parámetros
  - Sheet "CT_256slices"  — comparativa 3 fabricantes, 15 parámetros
  - Sheet "Resumen"       — tabla ejecutiva por segmento con recomendaciones

Uso:
    python build_comparison_excel.py

Output:
    outputs/MRI_CT_Technical_Comparison.xlsx
"""

import json
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ── Paleta de colores ─────────────────────────────────────────────────────────
SIEMENS_TEAL   = "009999"
GE_PURPLE      = "5B2D8E"
PHILIPS_BLUE   = "0B5394"
CANON_RED      = "CC0000"
HEADER_BG      = "1A3A5C"
SUBHEADER_BG   = "2E6DA4"
CATEGORY_BG    = "D6E4F0"
ALT_ROW        = "F0F7FF"
WHITE          = "FFFFFF"
LIGHT_GRAY     = "F5F5F5"
DARK_TEXT      = "1A1A1A"
MID_GRAY       = "CCCCCC"
BEST_GREEN     = "E2EFDA"
WORST_RED      = "FCE4D6"

BRAND_COLORS = {
    "Siemens": SIEMENS_TEAL,
    "GE":      GE_PURPLE,
    "Philips": PHILIPS_BLUE,
    "Canon":   CANON_RED,
}

thin = Side(style="thin", color=MID_GRAY)
thick = Side(style="medium", color="999999")
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)

# ── Helpers de estilo ─────────────────────────────────────────────────────────
def style_header(cell, bg=HEADER_BG, text_color=WHITE, size=11, bold=True, wrap=True):
    cell.font = Font(bold=bold, color=text_color, size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=wrap)
    cell.border = border_thin

def style_cell(cell, bg=WHITE, bold=False, color=DARK_TEXT, align="left",
               size=10, wrap=False, number_fmt=None):
    cell.font = Font(bold=bold, color=color, size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
    cell.border = border_thin
    if number_fmt:
        cell.number_format = number_fmt

def style_category(cell, text):
    cell.value = text
    cell.font = Font(bold=True, color=DARK_TEXT, size=10, name="Calibri",
                     italic=True)
    cell.fill = PatternFill("solid", fgColor=CATEGORY_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                                indent=1)
    cell.border = border_thin


# ── Carga de datos ────────────────────────────────────────────────────────────
with open("data/specs.json", encoding="utf-8") as f:
    data = json.load(f)

os.makedirs("outputs", exist_ok=True)
wb = Workbook()
wb.remove(wb.active)  # eliminar hoja por defecto


# ── Función principal: escribir hoja de comparativa ──────────────────────────
def write_comparison_sheet(wb, sheet_name, equipment_key, title):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    eq_data   = data["equipment"][equipment_key]
    segment   = eq_data["segment"]
    params    = data["parameters"][segment]
    models    = eq_data["models"]
    brands    = list(models.keys())

    # ── Título ──────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=3 + len(brands))
    tc = ws.cell(1, 1, title)
    tc.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    tc.fill = PatternFill("solid", fgColor=HEADER_BG)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    tc.border = border_thin
    ws.row_dimensions[1].height = 32

    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=3 + len(brands))
    sub = ws.cell(2, 1,
        "Fuente: datasheets oficiales fabricantes · Valores orientativos — "
        "verificar con fabricante para configuración específica")
    sub.font = Font(italic=True, color="666666", size=9, name="Calibri")
    sub.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Cabecera de fabricantes ──────────────────────────────────────────────
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 40

    # Columnas fijas
    style_header(ws.cell(3, 1), bg=SUBHEADER_BG)
    style_header(ws.cell(3, 2), bg=SUBHEADER_BG)
    style_header(ws.cell(3, 3), bg=SUBHEADER_BG)
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
    ws.cell(3, 1).value = "Categoría"
    ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
    ws.cell(3, 2).value = "Parámetro"
    ws.merge_cells(start_row=3, start_column=3, end_row=4, end_column=3)
    ws.cell(3, 3).value = "Unidad"

    for j, brand in enumerate(brands):
        col = 4 + j
        m = models[brand]
        # Fila marca
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col)
        bc = ws.cell(3, col, brand)
        style_header(bc, bg=BRAND_COLORS[brand], size=11)

        # Fila modelo
        mc = ws.cell(4, col, m["model"])
        style_header(mc, bg=BRAND_COLORS[brand], size=9, bold=False)

    # Anchos de columnas
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 12
    for j in range(len(brands)):
        ws.column_dimensions[get_column_letter(4 + j)].width = 26

    # ── Filas de parámetros ──────────────────────────────────────────────────
    current_category = None
    row = 5

    for param in params:
        cat   = param["category"]
        label = param["label"]
        unit  = param["unit"]
        pid   = param["id"]

        # Separador de categoría
        if cat != current_category:
            current_category = cat
            ws.row_dimensions[row].height = 18
            style_category(ws.cell(row, 1), cat)
            for col in range(2, 4 + len(brands)):
                c = ws.cell(row, col)
                c.fill = PatternFill("solid", fgColor=CATEGORY_BG)
                c.border = border_thin
            row += 1

        ws.row_dimensions[row].height = 22
        bg = ALT_ROW if row % 2 == 0 else WHITE

        # Categoría (sin texto — ya se mostró en separador)
        style_cell(ws.cell(row, 1), bg=bg, color="999999", size=9,
                   align="center")
        # Parámetro
        style_cell(ws.cell(row, 2, label), bg=bg, bold=False, size=10)
        ws.cell(row, 2).alignment = Alignment(horizontal="left",
                                               vertical="center", indent=1)
        ws.cell(row, 2).border = border_thin
        # Unidad
        style_cell(ws.cell(row, 3, unit), bg=bg, color="666666", size=9,
                   align="center")

        # Valores por fabricante
        values = []
        for brand in brands:
            v = models[brand].get(pid, "N/D")
            values.append(v)

        # Detectar mejor/peor valor numérico (solo para campos comparables)
        numeric_vals = []
        for v in values:
            try:
                numeric_vals.append(float(str(v).replace(",", ".")))
            except:
                numeric_vals.append(None)

        has_numeric = all(v is not None for v in numeric_vals)

        # Parámetros donde mayor = mejor
        higher_is_better = {
            "gradient_strength", "slew_rate", "bore_diameter", "patient_weight",
            "fov", "channels", "dose_reduction", "spatial_resolution",
            "detector_coverage", "tube_power"
        }
        # Parámetros donde menor = mejor
        lower_is_better = {
            "noise_db", "helium_boiloff", "power_consumption",
            "rotation_speed", "temporal_resolution", "bore_length"
        }

        best_idx = worst_idx = None
        if has_numeric and pid in higher_is_better:
            best_idx  = numeric_vals.index(max(numeric_vals))
            worst_idx = numeric_vals.index(min(numeric_vals))
        elif has_numeric and pid in lower_is_better:
            best_idx  = numeric_vals.index(min(numeric_vals))
            worst_idx = numeric_vals.index(max(numeric_vals))

        for j, (brand, val) in enumerate(zip(brands, values)):
            col = 4 + j
            cell = ws.cell(row, col, val)

            cell_bg = bg
            if j == best_idx:
                cell_bg = BEST_GREEN
            elif j == worst_idx:
                cell_bg = WORST_RED

            style_cell(cell, bg=cell_bg, align="center", size=10)

            # Negrita para Siemens
            if brand == "Siemens":
                cell.font = Font(bold=True, color=DARK_TEXT, size=10,
                                 name="Calibri")

        row += 1

    # ── Fila de highlights ────────────────────────────────────────────────────
    row += 1
    ws.row_dimensions[row].height = 18
    hl_title = ws.cell(row, 1, "VENTAJAS CLAVE")
    hl_title.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
    hl_title.fill = PatternFill("solid", fgColor=HEADER_BG)
    hl_title.alignment = Alignment(horizontal="center", vertical="center")
    hl_title.border = border_thin
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=3)
    ws.cell(row, 2).fill = PatternFill("solid", fgColor=HEADER_BG)
    ws.cell(row, 2).border = border_thin
    ws.cell(row, 3).fill = PatternFill("solid", fgColor=HEADER_BG)
    ws.cell(row, 3).border = border_thin

    for j, brand in enumerate(brands):
        col = 4 + j
        hc = ws.cell(row, col)
        style_header(hc, bg=BRAND_COLORS[brand], size=9)

    row += 1

    # Máx. 3 highlights por fabricante
    max_hl = max(len(models[b].get("highlights", [])) for b in brands)
    for i in range(min(3, max_hl)):
        ws.row_dimensions[row].height = 30
        style_cell(ws.cell(row, 1), bg=LIGHT_GRAY)
        style_cell(ws.cell(row, 2), bg=LIGHT_GRAY)
        style_cell(ws.cell(row, 3), bg=LIGHT_GRAY)
        for j, brand in enumerate(brands):
            col = 4 + j
            hls = models[brand].get("highlights", [])
            txt = hls[i] if i < len(hls) else ""
            c = ws.cell(row, col, txt)
            style_cell(c, bg=ALT_ROW if row % 2 == 0 else WHITE,
                       align="center", size=9, wrap=True)
        row += 1

    # ── Fila de fortalezas clínicas ────────────────────────────────────────
    row += 1
    ws.row_dimensions[row].height = 18
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=3)
    fc = ws.cell(row, 1, "APLICACIONES CLÍNICAS PRINCIPALES")
    fc.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
    fc.fill = PatternFill("solid", fgColor=SUBHEADER_BG)
    fc.alignment = Alignment(horizontal="center", vertical="center")
    fc.border = border_thin
    ws.cell(row, 2).fill = PatternFill("solid", fgColor=SUBHEADER_BG)
    ws.cell(row, 2).border = border_thin
    ws.cell(row, 3).fill = PatternFill("solid", fgColor=SUBHEADER_BG)
    ws.cell(row, 3).border = border_thin

    for j, brand in enumerate(brands):
        col = 4 + j
        style_header(ws.cell(row, col), bg=BRAND_COLORS[brand], size=9)

    row += 1
    ws.row_dimensions[row].height = 50
    style_cell(ws.cell(row, 1), bg=LIGHT_GRAY)
    style_cell(ws.cell(row, 2), bg=LIGHT_GRAY)
    style_cell(ws.cell(row, 3), bg=LIGHT_GRAY)
    for j, brand in enumerate(brands):
        col = 4 + j
        strengths = " · ".join(models[brand].get("clinical_strengths", []))
        c = ws.cell(row, col, strengths)
        style_cell(c, bg=ALT_ROW, align="center", size=9, wrap=True)

    return ws


# ── Crear hojas ───────────────────────────────────────────────────────────────
write_comparison_sheet(wb, "MRI_1.5T",
    "MRI_1.5T",
    "Comparativa Técnica — MRI 1.5T | Siemens MAGNETOM Sola · GE SIGNA Artist · Philips Ingenia 1.5T · Canon Vantage Orian")

write_comparison_sheet(wb, "MRI_3T",
    "MRI_3T",
    "Comparativa Técnica — MRI 3T | Siemens MAGNETOM Vida · GE SIGNA Premier · Philips Ingenia Elition 3T")

write_comparison_sheet(wb, "CT_64slices",
    "CT_64",
    "Comparativa Técnica — CT 64 Cortes | Siemens SOMATOM go.Up · GE Revolution EVO · Philips Incisive CT · Canon Aquilion Lightning")

write_comparison_sheet(wb, "CT_128slices",
    "CT_128",
    "Comparativa Técnica — CT 128 Cortes | Siemens SOMATOM go.Top · GE Revolution Ascend · Philips Incisive CT 128")

write_comparison_sheet(wb, "CT_256slices",
    "CT_256",
    "Comparativa Técnica — CT 256 Cortes | Siemens SOMATOM Definition AS+ · GE Revolution HD · Philips IQon Spectral CT")


# ── Hoja Resumen Ejecutivo ────────────────────────────────────────────────────
ws_res = wb.create_sheet("Resumen_Ejecutivo", 0)
ws_res.sheet_view.showGridLines = False

# Título
ws_res.merge_cells("B2:J2")
t = ws_res["B2"]
t.value = "Resumen Ejecutivo — Comparativa MRI & CT | Siemens vs GE vs Philips vs Canon"
t.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
t.fill = PatternFill("solid", fgColor=HEADER_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
t.border = border_thin
ws_res.row_dimensions[2].height = 32

ws_res.merge_cells("B3:J3")
s = ws_res["B3"]
s.value = ("Parámetros clave para evaluación en licitaciones públicas "
           "(SEACE / SERCOP) y procesos de adquisición privada — Perú & Ecuador")
s.font = Font(italic=True, color="555555", size=10, name="Calibri")
s.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
s.alignment = Alignment(horizontal="center", vertical="center")
s.border = border_thin
ws_res.row_dimensions[3].height = 20

summary_data = [
    # (Segmento, Equipo, Siemens modelo, GE modelo, Philips modelo, Canon modelo,
    #  Precio ref. USD, Recomendación de uso)
    ("MRI 1.5T",
     "Equipo estándar hospitalario\nReferente MINSA / EsSalud",
     "MAGNETOM Sola\n$850K–1.2M",
     "SIGNA Artist\n$800K–1.1M",
     "Ingenia 1.5T\n$820K–1.15M",
     "Vantage Orian\n$750K–1.05M",
     "Neurología, MSK, Oncología rutinaria\nIdeal para hospitales de mediana complejidad"),
    ("MRI 3T",
     "Equipo avanzado / investigación\nReferente centros especializados",
     "MAGNETOM Vida\n$1.4M–2.2M",
     "SIGNA Premier\n$1.35M–2.1M",
     "Ingenia Elition 3T\n$1.38M–2.15M",
     "—",
     "Neurología avanzada, fMRI, Cardiología\nIdeal para centros de excelencia e investigación"),
    ("CT 64 cortes",
     "CT básico polivalente\nReferente MINSA / clínicas privadas",
     "SOMATOM go.Up\n$280K–420K",
     "Revolution EVO\n$270K–410K",
     "Incisive CT 64\n$275K–415K",
     "Aquilion Lightning\n$265K–400K",
     "Rutina clínica, urgencias, tórax/abdomen\nRelación costo-efectividad más alta"),
    ("CT 128 cortes",
     "CT intermedio / cardiología\nReferente hospitales de alta complejidad",
     "SOMATOM go.Top\n$450K–680K",
     "Revolution Ascend\n$440K–660K",
     "Incisive CT 128\n$445K–670K",
     "—",
     "Cardiología, oncología, trauma\nMejor balance rendimiento/precio para hospitales terciarios"),
    ("CT 256 cortes",
     "CT alta gama / dual energy\nReferente centros de referencia INEN / EsSalud",
     "Definition AS+\n$700K–950K",
     "Revolution HD\n$680K–930K",
     "IQon Spectral CT\n$720K–970K",
     "—",
     "Cardiología avanzada, oncología espectral\nMayor inversión — justificado en centros de alta demanda"),
]

headers = ["Segmento", "Perfil de Uso", "Siemens", "GE", "Philips", "Canon",
           "Aplicación Recomendada"]
col_widths = [14, 28, 22, 22, 22, 22, 36]
header_bgs = [HEADER_BG, SUBHEADER_BG,
              SIEMENS_TEAL, GE_PURPLE, PHILIPS_BLUE, CANON_RED,
              HEADER_BG]

start_row = 5
# Headers
ws_res.row_dimensions[start_row].height = 36
for j, (h, bg) in enumerate(zip(headers, header_bgs)):
    col = j + 2
    c = ws_res.cell(start_row, col, h)
    style_header(c, bg=bg, size=10)
    ws_res.column_dimensions[get_column_letter(col)].width = col_widths[j]

for i, row_data in enumerate(summary_data):
    r = start_row + 1 + i
    ws_res.row_dimensions[r].height = 56
    bg = ALT_ROW if i % 2 == 0 else WHITE

    for j, val in enumerate(row_data):
        col = j + 2
        c = ws_res.cell(r, col, val)
        is_brand = j in (2, 3, 4, 5)
        brand_color = [SIEMENS_TEAL, GE_PURPLE, PHILIPS_BLUE, CANON_RED][j - 2] if is_brand else None

        c.font = Font(bold=(j == 0), color=DARK_TEXT, size=9, name="Calibri")
        if is_brand and val != "—":
            c.font = Font(bold=True, color=brand_color, size=9, name="Calibri")
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
        c.border = border_thin

# Nota al pie
last_r = start_row + len(summary_data) + 2
ws_res.merge_cells(start_row=last_r, start_column=2,
                   end_row=last_r, end_column=8)
note = ws_res.cell(last_r, 2,
    "Verde = valor más favorable | Rojo = valor menos favorable | "
    "Precios son referenciales FOB y varían por configuración, accesorios e instalación. "
    "Fuente: datasheets oficiales fabricantes (2024–2025).")
note.font = Font(italic=True, color="777777", size=8, name="Calibri")
note.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
note.border = border_thin
ws_res.row_dimensions[last_r].height = 30

# ── Guardar ───────────────────────────────────────────────────────────────────
out_path = "outputs/MRI_CT_Technical_Comparison.xlsx"
wb.save(out_path)
print(f"✅ Excel guardado: {out_path}")
print(f"   Hojas: Resumen_Ejecutivo | MRI_1.5T | MRI_3T | CT_64slices | CT_128slices | CT_256slices")
